import os
import re
import logging
import threading
from datetime import datetime, timezone
from dataclasses import dataclass, field
from concurrent.futures import ThreadPoolExecutor, as_completed

import paramiko

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from dotenv import load_dotenv

load_dotenv()

# =====================================================
# Logging
# =====================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger("linux_patching_ssh")
_log_lock = threading.Lock()


def log_info(vm_name: str, msg: str):
    with _log_lock:
        log.info(f"[{vm_name}] {msg}")


def log_error(vm_name: str, msg: str):
    with _log_lock:
        log.error(f"[{vm_name}] {msg}")


# =====================================================
# Configuration
# =====================================================
class PatchConfig:
    # Each target needs enough info to open an SSH session directly —
    TARGETS = [
         {
            "name": os.getenv("SSH_VMNAME"),
            "host": os.getenv("SSH_HOST"),
            "port": os.getenv("SSH_PORT"),
            "username": os.getenv("SSH_USERNAME"),
            "password": os.getenv("SSH_PASSWORD")
        }
    ]

    ASSESS_ONLY = False
    REBOOT_SETTING = "if_needed"  # never | if_needed

    SSH_CONNECT_TIMEOUT_SECONDS = 15
    COMMAND_TIMEOUT_SECONDS = 60 * 30  # 30 min hard ceiling per command
    MAX_WORKERS = 8

    REPORT_DIR = os.path.join(os.getcwd(), "reports")


@dataclass
class PatchResult:
    vm_name: str
    host: str = ""
    success: bool = False
    stage_reached: str = "Not started"
    distro_family: str = ""  # debian | rhel | suse | unknown
    os_name: str = ""
    os_version: str = ""
    reboot_status: str = ""  # Required / NotRequired / Rebooted / Unknown
    start_time: datetime = field(default_factory=lambda: datetime.now(timezone.utc))
    end_time: datetime = None
    duration_seconds: float = 0.0
    error: str = ""
    assess_raw_output: str = ""
    install_raw_output: str = ""
    packages_before: dict = field(default_factory=dict)
    packages_after: dict = field(default_factory=dict)
    package_version_error: str = ""
    version_changes: list = field(default_factory=list)


# =====================================================
# SSH connection + command execution helpers
# =====================================================
def _connect_ssh(target: dict, config: PatchConfig) -> paramiko.SSHClient:
    client = paramiko.SSHClient()
    client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    client.connect(
        hostname=target["host"],
        port=target.get("port", 22),
        username=target["username"],
        password=target["password"],
        timeout=config.SSH_CONNECT_TIMEOUT_SECONDS,
        banner_timeout=config.SSH_CONNECT_TIMEOUT_SECONDS,
        auth_timeout=config.SSH_CONNECT_TIMEOUT_SECONDS,
    )
    return client


def _run_ssh_command(ssh_client: paramiko.SSHClient, script_lines: list, config: PatchConfig, use_sudo: bool = True) -> str:
    """Runs a multi-line shell script over an existing SSH session and
    returns combined stdout+stderr. Raises TimeoutError if the command
    doesn't finish within COMMAND_TIMEOUT_SECONDS."""
    joined_script = "\n".join(script_lines)
    if use_sudo:
        # -S reads password from stdin if needed; most cloud images have
        # passwordless sudo configured for the provisioned user, so this
        # is mainly a safety net rather than the primary mechanism.
        command = f"sudo -n bash -c {paramiko_quote(joined_script)}"
    else:
        command = f"bash -c {paramiko_quote(joined_script)}"

    stdin, stdout, stderr = ssh_client.exec_command(command, timeout=config.COMMAND_TIMEOUT_SECONDS)
    out_text = stdout.read().decode(errors="replace")
    err_text = stderr.read().decode(errors="replace")
    return (out_text + "\n" + err_text).strip()


def paramiko_quote(text: str) -> str:
    """Wraps a multi-line script safely in single quotes for remote bash -c."""
    return "'" + text.replace("'", "'\\''") + "'"


# =====================================================
# Distro detection
# =====================================================
def _detect_distro_family(ssh_client, vm_name, config: PatchConfig):
    script = [
        "if [ -f /etc/os-release ]; then . /etc/os-release; "
        "echo \"OS_ID=${ID:-unknown}\"; echo \"OS_VERSION=${VERSION_ID:-unknown}\"; "
        "else echo OS_ID=unknown; echo OS_VERSION=unknown; fi",
        "if command -v apt-get >/dev/null 2>&1; then echo FAMILY=debian; "
        "elif command -v zypper >/dev/null 2>&1; then echo FAMILY=suse; "
        "elif command -v dnf >/dev/null 2>&1 || command -v yum >/dev/null 2>&1; then echo FAMILY=rhel; "
        "else echo FAMILY=unknown; fi",
    ]
    output = _run_ssh_command(ssh_client, script, config, use_sudo=False)

    family_match = re.search(r"FAMILY=(\S+)", output)
    os_id_match = re.search(r"OS_ID=(\S+)", output)
    os_version_match = re.search(r"OS_VERSION=(\S+)", output)

    distro_family = family_match.group(1) if family_match else "unknown"
    os_name = os_id_match.group(1).strip('"') if os_id_match else "unknown"
    os_version = os_version_match.group(1).strip('"') if os_version_match else "unknown"
    return distro_family, os_name, os_version


# =====================================================
# Package version snapshot
# =====================================================
_PACKAGE_DUMP_SCRIPT = (
    "dpkg-query -W -f='${Package}::${Version}\\n' 2>/dev/null "
    "|| rpm -qa --qf '%{NAME}::%{VERSION}-%{RELEASE}\\n' 2>/dev/null"
)


def _parse_package_dump(raw_output: str) -> dict:
    packages = {}
    for line in raw_output.splitlines():
        line = line.strip()
        if "::" not in line:
            continue
        name, _, version = line.partition("::")
        if name:
            packages[name] = version
    return packages


def _capture_package_versions(ssh_client, vm_name, label, config, result: PatchResult = None) -> dict:
    try:
        output = _run_ssh_command(ssh_client, [_PACKAGE_DUMP_SCRIPT], config, use_sudo=False)
        packages = _parse_package_dump(output)
        log_info(vm_name, f"{label}: captured {len(packages)} package versions")
        return packages
    except Exception as exc:
        log_error(vm_name, f"{label}: failed to capture package versions — {exc}")
        if result is not None:
            result.package_version_error = f"{label} capture failed: {exc}"
        return {}


def _diff_package_versions(before: dict, after: dict) -> list:
    changes = []
    for name, before_version in before.items():
        after_version = after.get(name)
        if after_version is None:
            changes.append({"name": name, "version_before": before_version, "version_after": "", "status": "Removed"})
        elif after_version != before_version:
            changes.append({"name": name, "version_before": before_version, "version_after": after_version, "status": "Updated"})
    for name, after_version in after.items():
        if name not in before:
            changes.append({"name": name, "version_before": "", "version_after": after_version, "status": "Added"})
    return changes


# =====================================================
# Assess / install scripts per distro family
# =====================================================
def _build_assess_script(distro_family: str) -> list:
    if distro_family == "debian":
        return [
            "export DEBIAN_FRONTEND=noninteractive",
            "apt-get update -y",
            "apt list --upgradable 2>/dev/null",
        ]
    if distro_family == "rhel":
        return [
            "(command -v dnf >/dev/null 2>&1 && dnf check-update) || yum check-update",
        ]
    if distro_family == "suse":
        return [
            "zypper refresh",
            "zypper list-updates",
        ]
    return ["echo 'UNSUPPORTED_DISTRO'"]


def _build_install_script(distro_family: str) -> list:
    if distro_family == "debian":
        return [
            "export DEBIAN_FRONTEND=noninteractive",
            "apt-get update -y",
            "apt-get upgrade -y",
            "apt-get autoremove -y",
            "if [ -f /var/run/reboot-required ]; then echo REBOOT_REQUIRED=yes; else echo REBOOT_REQUIRED=no; fi",
        ]
    if distro_family == "rhel":
        return [
            "(command -v dnf >/dev/null 2>&1 && dnf update -y) || yum update -y",
            "if command -v needs-restarting >/dev/null 2>&1; then "
            "needs-restarting -r >/dev/null 2>&1 && echo REBOOT_REQUIRED=no || echo REBOOT_REQUIRED=yes; "
            "else echo REBOOT_REQUIRED=unknown; fi",
        ]
    if distro_family == "suse":
        return [
            "zypper refresh",
            "zypper --non-interactive update",
            "if [ -f /var/run/reboot-needed ] || [ -f /boot/do_purge_kernels ]; then "
            "echo REBOOT_REQUIRED=yes; else echo REBOOT_REQUIRED=no; fi",
        ]
    return ["echo 'UNSUPPORTED_DISTRO'"]


# =====================================================
# Reboot handling — separate SSH session so the connection drop caused by
# a reboot doesn't get mistaken for a command failure.
# =====================================================
def _reboot_and_wait(target: dict, config: PatchConfig, vm_name: str, max_wait_seconds: int = 600):
    """Issues a reboot, then polls SSH connectivity until the VM comes
    back up or max_wait_seconds elapses."""
    import socket
    import time

    try:
        ssh_client = _connect_ssh(target, config)
        ssh_client.exec_command("sudo -n reboot")
        ssh_client.close()
    except Exception as exc:
        # A dropped connection here is expected — the VM is rebooting.
        log_info(vm_name, f"Reboot command sent (connection drop expected): {exc}")

    log_info(vm_name, "Waiting for VM to come back online after reboot...")
    time.sleep(15)  # give it a moment to actually go down before polling
    waited = 15
    while waited < max_wait_seconds:
        try:
            test_client = _connect_ssh(target, config)
            test_client.close()
            log_info(vm_name, f"VM is back online after ~{waited}s")
            return True
        except (socket.error, paramiko.SSHException, EOFError):
            time.sleep(10)
            waited += 10
    log_error(vm_name, f"VM did not come back online within {max_wait_seconds}s of reboot")
    return False


# =====================================================
# Per-VM patch job
# =====================================================
def patch_single_vm(target: dict, config: PatchConfig) -> PatchResult:
    vm_name = target["vm_name"]
    result = PatchResult(vm_name=vm_name, host=target["host"])
    ssh_client = None
    try:
        result.stage_reached = "Connecting via SSH"
        ssh_client = _connect_ssh(target, config)
        log_info(vm_name, f"SSH connected to {target['host']}")

        result.stage_reached = "Distro detection"
        result.distro_family, result.os_name, result.os_version = _detect_distro_family(ssh_client, vm_name, config)
        log_info(vm_name, f"Distro family: {result.distro_family} ({result.os_name} {result.os_version})")

        if result.distro_family == "unknown":
            result.error = "Could not detect apt-get, yum/dnf, or zypper on the VM"
            result.stage_reached = "Distro detection failed"
            return result

        result.stage_reached = "Capturing pre-patch package versions"
        result.packages_before = _capture_package_versions(ssh_client, vm_name, "Before-patch snapshot", config, result)

        result.stage_reached = "Assessment"
        assess_script = _build_assess_script(result.distro_family)
        result.assess_raw_output = _run_ssh_command(ssh_client, assess_script, config, use_sudo=True)
        log_info(vm_name, "Assessment complete")

        if config.ASSESS_ONLY:
            result.stage_reached = "Assess-only complete (install skipped)"
            result.success = True
            return result

        result.stage_reached = "Install"
        install_script = _build_install_script(result.distro_family)
        result.install_raw_output = _run_ssh_command(ssh_client, install_script, config, use_sudo=True)
        log_info(vm_name, "Install complete")

        if "REBOOT_REQUIRED=yes" in result.install_raw_output:
            result.reboot_status = "Required"
        elif "REBOOT_REQUIRED=no" in result.install_raw_output:
            result.reboot_status = "NotRequired"
        else:
            result.reboot_status = "Unknown"

        if result.reboot_status == "Required" and config.REBOOT_SETTING == "if_needed":
            log_info(vm_name, "Reboot required — rebooting and waiting for VM to come back")
            ssh_client.close()
            ssh_client = None
            came_back = _reboot_and_wait(target, config, vm_name)
            result.reboot_status = "Rebooted" if came_back else "Rebooted (did not confirm back online)"
            ssh_client = _connect_ssh(target, config)

        result.stage_reached = "Capturing post-patch package versions"
        result.packages_after = _capture_package_versions(ssh_client, vm_name, "After-patch snapshot", config, result)
        result.version_changes = _diff_package_versions(result.packages_before, result.packages_after)

        result.stage_reached = "Completed"
        result.success = True

    except paramiko.AuthenticationException as exc:
        result.error = f"SSH authentication failed: {exc}"
        log_error(vm_name, result.error)
    except (paramiko.SSHException, OSError) as exc:
        result.error = f"SSH connection error: {exc}"
        log_error(vm_name, result.error)
    except Exception as exc:
        result.error = f"Unexpected error: {exc}"
        log_error(vm_name, result.error)
    finally:
        if ssh_client is not None:
            try:
                ssh_client.close()
            except Exception:
                pass
        result.end_time = datetime.now(timezone.utc)
        result.duration_seconds = round((result.end_time - result.start_time).total_seconds(), 1)

    return result


# =====================================================
# Job runner — parallel across targets
# =====================================================
def run_patching_job(config: PatchConfig):
    if not config.TARGETS:
        log.error("PatchConfig.TARGETS is empty — add VM connection details before running.")
        return []

    results = []
    with ThreadPoolExecutor(max_workers=config.MAX_WORKERS) as executor:
        futures = {executor.submit(patch_single_vm, target, config): target for target in config.TARGETS}
        for future in as_completed(futures):
            target = futures[future]
            try:
                results.append(future.result())
            except Exception as exc:
                log_error(target["vm_name"], f"Unhandled exception in worker: {exc}")
                results.append(PatchResult(vm_name=target["vm_name"], host=target.get("host", ""), error=str(exc)))
    return results


# =====================================================
# Excel report
# =====================================================
def write_excel_report(results, config: PatchConfig) -> str:
    os.makedirs(config.REPORT_DIR, exist_ok=True)
    filename = f"linux_ssh_patch_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(config.REPORT_DIR, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="4472C4")

    headers = [
        "VM Name", "Host", "OS", "OS Version", "Package Manager", "Success", "Stage Reached",
        "Reboot Status", "Packages Changed", "Error", "Duration (s)",
    ]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in results:
        ws.append([
            r.vm_name, r.host, r.os_name, r.os_version, r.distro_family,
            "Yes" if r.success else "No", r.stage_reached,
            r.reboot_status, len(r.version_changes), r.error, r.duration_seconds,
        ])

    for col_idx, header in enumerate(headers, start=1):
        max_len = max(len(header), *(len(str(ws.cell(row=row, column=col_idx).value or "")) for row in range(2, ws.max_row + 1))) if ws.max_row > 1 else len(header)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    ws2 = wb.create_sheet("Version Changes")
    ws2.append(["VM Name", "Package", "Before", "After", "Status"])
    for col_idx in range(1, 6):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
    for r in results:
        for ch in r.version_changes:
            ws2.append([r.vm_name, ch["name"], ch["version_before"], ch["version_after"], ch["status"]])
    ws2.freeze_panes = "A2"

    ws3 = wb.create_sheet("VM Detail Log")
    ws3.append(["VM Name", "Host", "Summary", "Assess Output (raw)", "Install Output (raw)"])
    for col_idx in range(1, 6):
        cell = ws3.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    for r in results:
        summary_lines = [
            f"OS: {r.os_name} {r.os_version} ({r.distro_family})",
            f"Status: {'Success' if r.success else 'Failed'} (stage reached: {r.stage_reached})",
            f"Reboot status: {r.reboot_status}",
            f"Packages before={len(r.packages_before)}, after={len(r.packages_after)}, changed={len(r.version_changes)}",
        ]
        if r.package_version_error:
            summary_lines.append(f"Package snapshot error: {r.package_version_error}")
        if r.error:
            summary_lines.append(f"Error: {r.error}")
        summary_lines.append(f"Duration: {r.duration_seconds}s")

        ws3.append([r.vm_name, r.host, "\n".join(summary_lines), r.assess_raw_output.strip(), r.install_raw_output.strip()])
        row_idx = ws3.max_row
        for col_idx in (1, 2):
            ws3.cell(row=row_idx, column=col_idx).font = Font(name="Arial")
        for col_idx in (3, 4, 5):
            cell = ws3.cell(row=row_idx, column=col_idx)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws3.row_dimensions[row_idx].height = max(60, 13 * (len(summary_lines) + 4))

    ws3.column_dimensions["A"].width = 18
    ws3.column_dimensions["B"].width = 16
    ws3.column_dimensions["C"].width = 55
    ws3.column_dimensions["D"].width = 60
    ws3.column_dimensions["E"].width = 60
    ws3.freeze_panes = "A2"

    ws4 = wb.create_sheet("Run Info")
    ws4["A1"] = "Patching Run Summary"
    ws4["A1"].font = Font(name="Arial", bold=True, size=14)
    info_rows = [
        ("Report generated (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")),
        ("Total VMs targeted", len(results)),
        ("Succeeded", sum(1 for r in results if r.success)),
        ("Failed", sum(1 for r in results if not r.success)),
        ("Assess only (no install)", config.ASSESS_ONLY),
        ("Reboot setting", config.REBOOT_SETTING),
        ("Max parallel workers", config.MAX_WORKERS),
    ]
    for i, (label, value) in enumerate(info_rows, start=3):
        ws4.cell(row=i, column=1, value=label).font = Font(name="Arial", bold=True)
        ws4.cell(row=i, column=2, value=value).font = Font(name="Arial")
    ws4.column_dimensions["A"].width = 28
    ws4.column_dimensions["B"].width = 30

    wb.save(filepath)
    log.info(f"Excel report written to {filepath}")
    return filepath


if __name__ == "__main__":
    cfg = PatchConfig()
    job_results = run_patching_job(cfg)
    report_path = write_excel_report(job_results, cfg)
    print(f"\nReport: {report_path}")