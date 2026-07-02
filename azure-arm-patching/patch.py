import os
import sys
import re
import logging
import threading
from datetime import datetime, timezone
from dataclasses import dataclass, field
from concurrent.futures import ThreadPoolExecutor, as_completed
from dotenv import load_dotenv


from azure.identity import ClientSecretCredential
from azure.mgmt.compute import ComputeManagementClient
from azure.mgmt.compute.models import RunCommandInput, RunCommandInputParameter
from azure.core.exceptions import HttpResponseError, ResourceNotFoundError

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


load_dotenv()

# =====================================================
# Logging
# =====================================================
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
)
log = logging.getLogger("linux_patching_runcmd")
_log_lock = threading.Lock()


def log_info(vm_name: str, msg: str):
    with _log_lock:
        log.info(f"[{vm_name}] {msg}")


def log_error(vm_name: str, msg: str):
    with _log_lock:
        log.error(f"[{vm_name}] {msg}")


# =====================================================
# Configuration
# =====================================================
class PatchConfig:
    TARGETS = [
        ("Resource_group_name", "VM_name"),
        ("Resource_group_name", "VM_name"),
    ]

    # Whether to only assess (dry-run) and skip the actual install.
    # Strongly recommended to run once with this True before ever
    # flipping it to False against a fleet.
    ASSESS_ONLY = False

    # Reboot if the package manager / a kernel update says it's needed.
    # "never" = don't reboot even if needed. "if_needed" = reboot only
    # if reboot-required marker is present after install.
    REBOOT_SETTING = "if_needed"  # never | if_needed

    RUN_COMMAND_TIMEOUT_SECONDS = 60 * 60  # 1 hour hard ceiling per Run Command call
    MAX_WORKERS = 8

    REPORT_DIR = os.path.join(os.getcwd(), "reports")


@dataclass
class PatchResult:
    resource_group: str
    vm_name: str
    success: bool = False
    stage_reached: str = "Not started"
    distro_family: str = ""  # "debian" or "rhel" or "suse" or "unknown"
    os_name: str = ""  # actual OS, e.g. "ubuntu", "debian", "rhel", "centos", "sles", "opensuse-leap"
    os_version: str = ""  # e.g. "22.04", "11", "15.4"
    reboot_status: str = ""  # Required / NotRequired / Rebooted / Unknown
    start_time: datetime = field(default_factory=lambda: datetime.now(timezone.utc))
    end_time: datetime = None
    duration_seconds: float = 0.0
    error: str = ""
    assess_raw_output: str = ""
    install_raw_output: str = ""
    packages_before: dict = field(default_factory=dict)
    packages_after: dict = field(default_factory=dict)
    package_version_error: str = ""
    version_changes: list = field(default_factory=list)


def get_credential(tenant_id: str, client_id: str, client_secret: str) -> ClientSecretCredential:
    if not all([tenant_id, client_id, client_secret]):
        log.error("Missing Azure credentials — tenant_id, client_id, and client_secret are required.")
        sys.exit(1)
    return ClientSecretCredential(tenant_id=tenant_id, client_id=client_id, client_secret=client_secret)


# =====================================================
# Run Command helper — synchronous wrapper around begin_run_command
# =====================================================
def _run_command(
    client: ComputeManagementClient,
    resource_group: str,
    vm_name: str,
    script_lines: list,
    config: PatchConfig,
) -> str:
    """Executes a shell script on the VM via Azure Run Command (RunShellScript)
    and returns the combined stdout/stderr message text. Raises on Run
    Command failure or timeout."""
    run_command_input = RunCommandInput(
        command_id="RunShellScript",
        script=script_lines,
    )
    poller = client.virtual_machines.begin_run_command(
        resource_group_name=resource_group,
        vm_name=vm_name,
        parameters=run_command_input,
    )
    result = poller.result(timeout=config.RUN_COMMAND_TIMEOUT_SECONDS)

    output_text = ""
    for value in getattr(result, "value", []) or []:
        msg = getattr(value, "message", "") or ""
        output_text += msg + "\n"
    return output_text


# =====================================================
# Distro detection
# =====================================================
def _detect_distro_family(client, resource_group, vm_name, config: PatchConfig):
    """Returns (distro_family, os_name, os_version).

    distro_family is which package manager to use: "debian" (apt-get),
    "rhel" (yum/dnf), "suse" (zypper), or "unknown".
    os_name/os_version come from /etc/os-release so e.g. Debian and
    Ubuntu (both "debian" family) are still distinguishable in reports.
    """
    script = [
        "if [ -f /etc/os-release ]; then . /etc/os-release; "
        "echo \"OS_ID=${ID:-unknown}\"; echo \"OS_VERSION=${VERSION_ID:-unknown}\"; "
        "else echo OS_ID=unknown; echo OS_VERSION=unknown; fi",
        "if command -v apt-get >/dev/null 2>&1; then echo FAMILY=debian; "
        "elif command -v zypper >/dev/null 2>&1; then echo FAMILY=suse; "
        "elif command -v dnf >/dev/null 2>&1 || command -v yum >/dev/null 2>&1; then echo FAMILY=rhel; "
        "else echo FAMILY=unknown; fi",
    ]
    output = _run_command(client, resource_group, vm_name, script, config)

    family_match = re.search(r"FAMILY=(\S+)", output)
    os_id_match = re.search(r"OS_ID=(\S+)", output)
    os_version_match = re.search(r"OS_VERSION=(\S+)", output)

    distro_family = family_match.group(1) if family_match else "unknown"
    os_name = os_id_match.group(1).strip('"') if os_id_match else "unknown"
    os_version = os_version_match.group(1).strip('"') if os_version_match else "unknown"
    return distro_family, os_name, os_version


# =====================================================
# Package version snapshot (same idea as the Guest Patching version)
# =====================================================
_PACKAGE_DUMP_SCRIPT = (
    "dpkg-query -W -f='${Package}::${Version}\\n' 2>/dev/null "
    "|| rpm -qa --qf '%{NAME}::%{VERSION}-%{RELEASE}\\n' 2>/dev/null"
)


def _parse_package_dump(raw_output: str) -> dict:
    packages = {}
    for line in raw_output.splitlines():
        line = line.strip()
        if "::" not in line:
            continue
        name, _, version = line.partition("::")
        if name:
            packages[name] = version
    return packages


def _capture_package_versions(client, resource_group, vm_name, label, config, result: PatchResult = None) -> dict:
    try:
        output = _run_command(client, resource_group, vm_name, [_PACKAGE_DUMP_SCRIPT], config)
        packages = _parse_package_dump(output)
        log_info(vm_name, f"{label}: captured {len(packages)} package versions")
        return packages
    except Exception as exc:
        log_error(vm_name, f"{label}: failed to capture package versions — {exc}")
        if result is not None:
            result.package_version_error = f"{label} capture failed: {exc}"
        return {}


def _diff_package_versions(before: dict, after: dict) -> list:
    changes = []
    for name, before_version in before.items():
        after_version = after.get(name)
        if after_version is None:
            changes.append({"name": name, "version_before": before_version, "version_after": "", "status": "Removed"})
        elif after_version != before_version:
            changes.append({"name": name, "version_before": before_version, "version_after": after_version, "status": "Updated"})
    for name, after_version in after.items():
        if name not in before:
            changes.append({"name": name, "version_before": "", "version_after": after_version, "status": "Added"})
    return changes


# =====================================================
# Assess / install scripts per distro family
# =====================================================
def _build_assess_script(distro_family: str) -> list:
    if distro_family == "debian":
        return [
            "export DEBIAN_FRONTEND=noninteractive",
            "apt-get update -y",
            "apt list --upgradable 2>/dev/null",
        ]
    if distro_family == "rhel":
        return [
            "(command -v dnf >/dev/null 2>&1 && dnf check-update) || yum check-update",
        ]
    if distro_family == "suse":
        return [
            "zypper refresh",
            "zypper list-updates",
        ]
    return ["echo 'UNSUPPORTED_DISTRO'"]


def _build_install_script(distro_family: str, reboot_setting: str) -> list:
    if distro_family == "debian":
        lines = [
            "export DEBIAN_FRONTEND=noninteractive",
            "apt-get update -y",
            "apt-get upgrade -y",
            "apt-get autoremove -y",
            "if [ -f /var/run/reboot-required ]; then echo REBOOT_REQUIRED=yes; else echo REBOOT_REQUIRED=no; fi",
        ]
    elif distro_family == "rhel":
        lines = [
            "(command -v dnf >/dev/null 2>&1 && dnf update -y) || yum update -y",
            "if command -v needs-restarting >/dev/null 2>&1; then "
            "needs-restarting -r >/dev/null 2>&1 && echo REBOOT_REQUIRED=no || echo REBOOT_REQUIRED=yes; "
            "else echo REBOOT_REQUIRED=unknown; fi",
        ]
    elif distro_family == "suse":
        lines = [
            "zypper refresh",
            "zypper --non-interactive update",
            # zypper's own exit-code convention: 102 means update succeeded
            # but a reboot is needed. We can't see the exit code after the
            # fact from Run Command output alone, so fall back to the
            # reboot-needed indicator file used by zypper/SUSE tooling.
            "if [ -f /var/run/reboot-needed ] || [ -f /boot/do_purge_kernels ]; then "
            "echo REBOOT_REQUIRED=yes; else echo REBOOT_REQUIRED=no; fi",
        ]
    else:
        return ["echo 'UNSUPPORTED_DISTRO'"]

    if reboot_setting == "if_needed":
        lines.append(
            "grep -q 'REBOOT_REQUIRED=yes' <<< \"$(tail -n1 <<< \\\"$0\\\")\" || true"
        )
        # Actual reboot trigger is handled by the caller after parsing output,
        # not inline here — rebooting mid-script would cut off the Run Command result.
    return lines


# =====================================================
# Per-VM patch job
# =====================================================
def patch_single_vm(
    client: ComputeManagementClient,
    resource_group: str,
    vm_name: str,
    config: PatchConfig,
) -> PatchResult:
    result = PatchResult(resource_group=resource_group, vm_name=vm_name)
    try:
        log_info(vm_name, "Detecting distro family")
        result.stage_reached = "Distro detection"
        result.distro_family, result.os_name, result.os_version = _detect_distro_family(
            client, resource_group, vm_name, config
        )
        log_info(vm_name, f"Distro family: {result.distro_family} ({result.os_name} {result.os_version})")

        if result.distro_family == "unknown":
            result.error = "Could not detect apt-get, yum/dnf, or zypper on the VM"
            result.stage_reached = "Distro detection failed"
            return result

        result.stage_reached = "Capturing pre-patch package versions"
        result.packages_before = _capture_package_versions(
            client, resource_group, vm_name, "Before-patch snapshot", config, result
        )

        result.stage_reached = "Assessment"
        assess_script = _build_assess_script(result.distro_family)
        result.assess_raw_output = _run_command(client, resource_group, vm_name, assess_script, config)
        log_info(vm_name, "Assessment complete")

        if config.ASSESS_ONLY:
            result.stage_reached = "Assess-only complete (install skipped)"
            result.success = True
            return result

        result.stage_reached = "Install"
        install_script = _build_install_script(result.distro_family, config.REBOOT_SETTING)
        result.install_raw_output = _run_command(client, resource_group, vm_name, install_script, config)
        log_info(vm_name, "Install complete")

        if "REBOOT_REQUIRED=yes" in result.install_raw_output:
            result.reboot_status = "Required"
        elif "REBOOT_REQUIRED=no" in result.install_raw_output:
            result.reboot_status = "NotRequired"
        else:
            result.reboot_status = "Unknown"

        if result.reboot_status == "Required" and config.REBOOT_SETTING == "if_needed":
            log_info(vm_name, "Reboot required — issuing restart")
            client.virtual_machines.begin_restart(resource_group, vm_name).result(
                timeout=config.RUN_COMMAND_TIMEOUT_SECONDS
            )
            result.reboot_status = "Rebooted"

        result.stage_reached = "Capturing post-patch package versions"
        result.packages_after = _capture_package_versions(
            client, resource_group, vm_name, "After-patch snapshot", config, result
        )
        result.version_changes = _diff_package_versions(result.packages_before, result.packages_after)

        result.stage_reached = "Completed"
        result.success = True

    except (HttpResponseError, ResourceNotFoundError) as exc:
        result.error = f"Azure API error: {exc}"
        log_error(vm_name, result.error)
    except TimeoutError as exc:
        result.error = f"Timed out: {exc}"
        log_error(vm_name, result.error)
    except Exception as exc:
        result.error = f"Unexpected error: {exc}"
        log_error(vm_name, result.error)
    finally:
        result.end_time = datetime.now(timezone.utc)
        result.duration_seconds = round((result.end_time - result.start_time).total_seconds(), 1)

    return result


# =====================================================
# Job runner — parallel across targets
# =====================================================
def run_patching_job(tenant_id, client_id, client_secret, subscription_id, config: PatchConfig):
    credential = get_credential(tenant_id, client_id, client_secret)
    client = ComputeManagementClient(credential, subscription_id)

    results = []
    with ThreadPoolExecutor(max_workers=config.MAX_WORKERS) as executor:
        futures = {
            executor.submit(patch_single_vm, client, rg, vm, config): (rg, vm)
            for rg, vm in config.TARGETS
        }
        for future in as_completed(futures):
            rg, vm = futures[future]
            try:
                results.append(future.result())
            except Exception as exc:
                log_error(vm, f"Unhandled exception in worker: {exc}")
                failed = PatchResult(resource_group=rg, vm_name=vm, error=str(exc))
                results.append(failed)
    return results


# =====================================================
# Minimal Excel report — Run Info, Results, Version Changes
# =====================================================
def write_excel_report(results, config: PatchConfig) -> str:
    os.makedirs(config.REPORT_DIR, exist_ok=True)
    filename = f"linux_runcmd_patch_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    filepath = os.path.join(config.REPORT_DIR, filename)

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    header_font = Font(name="Arial", bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="4472C4")

    headers = [
        "Resource Group", "VM Name", "OS", "OS Version", "Package Manager", "Success", "Stage Reached",
        "Reboot Status", "Packages Changed", "Error", "Duration (s)",
    ]
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in results:
        ws.append([
            r.resource_group, r.vm_name, r.os_name, r.os_version, r.distro_family,
            "Yes" if r.success else "No", r.stage_reached,
            r.reboot_status, len(r.version_changes), r.error, r.duration_seconds,
        ])

    for col_idx, header in enumerate(headers, start=1):
        max_len = max(len(header), *(len(str(ws.cell(row=row, column=col_idx).value or "")) for row in range(2, ws.max_row + 1))) if ws.max_row > 1 else len(header)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    ws2 = wb.create_sheet("Version Changes")
    ws2.append(["Resource Group", "VM Name", "Package", "Before", "After", "Status"])
    for col_idx in range(1, 7):
        cell = ws2.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
    for r in results:
        for ch in r.version_changes:
            ws2.append([r.resource_group, r.vm_name, ch["name"], ch["version_before"], ch["version_after"], ch["status"]])
    ws2.freeze_panes = "A2"

    # ---- Detailed per-VM log: raw assess output, raw install output, and a
    # human-readable summary block, so you can see exactly what ran on each VM. ----
    ws3 = wb.create_sheet("VM Detail Log")
    ws3.append(["Resource Group", "VM Name", "Summary", "Assess Output (raw)", "Install Output (raw)"])
    for col_idx in range(1, 6):
        cell = ws3.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for r in results:
        summary_lines = [
            f"OS: {r.os_name} {r.os_version} ({r.distro_family})",
            f"Status: {'Success' if r.success else 'Failed'} (stage reached: {r.stage_reached})",
            f"Reboot status: {r.reboot_status}",
            f"Packages before={len(r.packages_before)}, after={len(r.packages_after)}, changed={len(r.version_changes)}",
        ]
        if r.package_version_error:
            summary_lines.append(f"Package snapshot error: {r.package_version_error}")
        if r.error:
            summary_lines.append(f"Error: {r.error}")
        summary_lines.append(f"Duration: {r.duration_seconds}s")

        ws3.append([
            r.resource_group, r.vm_name, "\n".join(summary_lines),
            r.assess_raw_output.strip(), r.install_raw_output.strip(),
        ])
        row_idx = ws3.max_row
        for col_idx in (1, 2):
            ws3.cell(row=row_idx, column=col_idx).font = Font(name="Arial")
        for col_idx in (3, 4, 5):
            cell = ws3.cell(row=row_idx, column=col_idx)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws3.row_dimensions[row_idx].height = max(60, 13 * (len(summary_lines) + 4))

    ws3.column_dimensions["A"].width = 20
    ws3.column_dimensions["B"].width = 18
    ws3.column_dimensions["C"].width = 55
    ws3.column_dimensions["D"].width = 60
    ws3.column_dimensions["E"].width = 60
    ws3.freeze_panes = "A2"

    # ---- Run-level summary sheet ----
    ws4 = wb.create_sheet("Run Info")
    ws4["A1"] = "Patching Run Summary"
    ws4["A1"].font = Font(name="Arial", bold=True, size=14)
    info_rows = [
        ("Report generated (UTC)", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")),
        ("Total VMs targeted", len(results)),
        ("Succeeded", sum(1 for r in results if r.success)),
        ("Failed", sum(1 for r in results if not r.success)),
        ("Assess only (no install)", config.ASSESS_ONLY),
        ("Reboot setting", config.REBOOT_SETTING),
        ("Max parallel workers", config.MAX_WORKERS),
    ]
    for i, (label, value) in enumerate(info_rows, start=3):
        ws4.cell(row=i, column=1, value=label).font = Font(name="Arial", bold=True)
        ws4.cell(row=i, column=2, value=value).font = Font(name="Arial")
    ws4.column_dimensions["A"].width = 28
    ws4.column_dimensions["B"].width = 30

    wb.save(filepath)
    log.info(f"Excel report written to {filepath}")
    return filepath


if __name__ == "__main__":
    TENANT_ID = os.environ.get("AZURE_TENANT_ID")
    CLIENT_ID = os.environ.get("AZURE_CLIENT_ID")
    CLIENT_SECRET = os.environ.get("AZURE_CLIENT_SECRET")
    SUBSCRIPTION_ID = os.environ.get("AZURE_SUBSCRIPTION_ID")

    if not all([TENANT_ID, CLIENT_ID, CLIENT_SECRET, SUBSCRIPTION_ID]):
        log.error(
            "Set AZURE_TENANT_ID, AZURE_CLIENT_ID, AZURE_CLIENT_SECRET, "
            "AZURE_SUBSCRIPTION_ID environment variables before running."
        )
        sys.exit(1)

    cfg = PatchConfig()
    job_results = run_patching_job(TENANT_ID, CLIENT_ID, CLIENT_SECRET, SUBSCRIPTION_ID, cfg)
    report_path = write_excel_report(job_results, cfg)
    print(f"\nReport: {report_path}")